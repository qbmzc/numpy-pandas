import csv
import openpyxl
import os
# 在下方输入需要合并的文件所在文件夹位置
path = 'D:/temp/1/'

file_list = os.listdir(path)
file_list.sort()

# Main
file_name = ''
x1 = 1
x2 = 1
fileNum = len(file_list)
print("在该目录下有%d个xlsx文件" % fileNum)
for file in file_list:
    if '1' in file:  # 此处需要修改！
        file_name = os.path.join(path, file)
    else:
        continue
    print(file_name)
    wb = openpyxl.load_workbook(file_name)

    # 获取workbook中所有的表格
    sheets = wb.sheetnames

    print(sheets)

    total_num = 0

    list1 = []
    # 循环遍历所有sheet
    for t in sheets:
        sheet = wb[t]
        print(sheet.title + '->>>' + '开始合并')

        # len_row代表表中有多少行数据，len_column代表excel表中有多少列数据
        len_row = sheet.max_row
        print(sheet.title + '数据量为' + str(len_row) + '行')
        total_num = len_row + total_num
        len_column = sheet.max_column
        # 合并的时候只保留第一张表的表头部分
        if t == 0:
            for i in range(1, len_row + 1):
                list2 = []
                for j in range(1, len_column + 1):
                    list2.append(sheet.cell(row=i, column=j).value)
                list1.append(list2)
        else:
            for i in range(2, len_row + 1):
                list2 = []
                for j in range(1, len_column + 1):
                    list2.append(sheet.cell(row=i, column=j).value)
                list1.append(list2)
    # print(list1)
    print("合并后数据量" + str(total_num))
    with open('C:/Users/admin/Desktop/test3.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(list1)
        f.close()
